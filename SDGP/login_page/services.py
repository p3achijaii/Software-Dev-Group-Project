import base64
import re
from calendar import monthrange
from datetime import datetime, timedelta
from urllib.parse import urlencode

from django.contrib.auth import get_user_model
from django.db.models import Q
from django.urls import reverse
from django.utils.text import slugify
from django.utils import timezone

from .models import (
    PlatformAccount,
    PlatformDashboardActivity,
    PlatformDashboardQuickToolSlot,
    PlatformInboxMessage,
    PlatformPasswordHistory,
    PlatformProfile,
    PlatformReport,
    PlatformScheduleEvent,
    PlatformTeam,
    PlatformTeamDependency,
)

USER_HOME_QUICK_TOOL_SLOT_COUNT = 3
USER_HOME_ACTIVITY_VISIBLE_HOURS = 24
USER_HOME_NOTIFICATION_VISIBLE_COUNT = 3
EMAIL_ADDRESS_PATTERN = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
USER_HOME_QUICK_TOOLS_CATALOG = [
    {
        'id': 'tile-1',
        'title': 'Message',
        'description': 'Send a quick message',
        'url_name': 'user_home_tool_message',
    },
    {
        'id': 'tile-2',
        'title': 'Report',
        'description': 'Generate a report',
        'url_name': 'user_home_tool_report',
    },
    {
        'id': 'tile-3',
        'title': 'Organisation',
        'description': 'View organisation details',
        'url_name': 'user_home_tool_organisation',
    },
    {
        'id': 'tile-4',
        'title': 'Calendar',
        'description': 'View and manage calendar events',
        'url_name': 'user_home_tool_calendar',
    },
    {
        'id': 'tile-5',
        'title': 'Team',
        'description': 'View team members',
        'url_name': 'user_home_tool_team',
    },
    {
        'id': 'tile-6',
        'title': 'Data Visualisation',
        'description': 'Explore team and skill charts',
        'url_name': 'user_home_tool_data',
    },
]
USER_HOME_QUICK_TOOL_IDS = {toolItem['id'] for toolItem in USER_HOME_QUICK_TOOLS_CATALOG}
SCHEDULE_DAY_START_MINUTES = 7 * 60
SCHEDULE_DAY_END_MINUTES = 23 * 60
SCHEDULE_COLOR_PALETTE = [
    'rgba(239, 68, 68, 1)',
    'rgba(249, 115, 22, 1)',
    'rgba(234, 179, 8, 1)',
    'rgba(34, 197, 94, 1)',
    'rgba(20, 184, 166, 1)',
    'rgba(14, 165, 233, 1)',
    'rgba(37, 99, 235, 1)',
    'rgba(99, 102, 241, 1)',
    'rgba(168, 85, 247, 1)',
    'rgba(236, 72, 153, 1)',
]
SCHEDULE_EVENT_WEEKDAY_BLUEPRINT = {
    1: [
        {
            'title': 'Sprint Planning',
            'start_time': '09:00',
            'end_time': '10:00',
            'platform': 'Teams',
            'invite_members': 'Product Team',
            'color': 'rgba(14, 165, 233, 1)',
            'color_secondary': 'rgba(37, 99, 235, 1)',
        },
    ],
    2: [
        {
            'title': 'Client Check-In',
            'start_time': '11:00',
            'end_time': '12:00',
            'platform': 'Zoom',
            'invite_members': 'Client + PM',
            'color': 'rgba(34, 197, 94, 1)',
            'color_secondary': 'rgba(20, 184, 166, 1)',
        },
        {
            'title': 'Roadmap Review',
            'start_time': '15:00',
            'end_time': '16:00',
            'platform': 'Google Meet',
            'invite_members': 'Leads',
            'color': 'rgba(20, 184, 166, 1)',
            'color_secondary': 'rgba(14, 165, 233, 1)',
        },
    ],
    3: [
        {
            'title': 'Design Review',
            'start_time': '10:00',
            'end_time': '11:30',
            'platform': 'Google Meet',
            'invite_members': 'UX Team',
            'color': 'rgba(249, 115, 22, 1)',
            'color_secondary': 'rgba(234, 179, 8, 1)',
        },
    ],
    4: [
        {
            'title': 'QA Sync',
            'start_time': '13:00',
            'end_time': '14:00',
            'platform': 'Teams',
            'invite_members': 'QA + Dev',
            'color': 'rgba(99, 102, 241, 1)',
            'color_secondary': 'rgba(37, 99, 235, 1)',
        },
        {
            'title': 'Release Prep',
            'start_time': '16:00',
            'end_time': '17:00',
            'platform': 'Zoom',
            'invite_members': 'Ops Team',
            'color': 'rgba(168, 85, 247, 1)',
            'color_secondary': 'rgba(99, 102, 241, 1)',
        },
    ],
    5: [
        {
            'title': 'Weekly Retrospective',
            'start_time': '14:00',
            'end_time': '15:00',
            'platform': 'Google Hangouts',
            'invite_members': 'All Members',
            'color': 'rgba(236, 72, 153, 1)',
            'color_secondary': 'rgba(239, 68, 68, 1)',
        },
    ],
}
PLATFORM_TEAM_DEFAULT_SEED = [
    {
        'name': 'The Codebreakers',
        'team_lead': 'William Cooper',
        'department_name': 'xTV Web',
        'department_head': 'Sultan Suleyman',
        'github_link': 'https://github.com/sky/the-codebreakers',
        'status': PlatformTeam.TeamStatusChoices.ACTIVE,
        'git_project_name': 'Client Lightning Xtv',
        'jira_project_name': 'Client Lightning Xtv',
        'jira_board_link': 'https://jira.example.com/the-codebreakers',
        'wiki_link': 'https://wiki.example.com/the-codebreakers',
        'slack_link': 'https://slack.com',
        'daily_standup_time': '11:00',
        'about_team': 'The Codebreakers team owns delivery and support for customer-facing platform work.',
        'key_skills': 'JavaScript | Django | Git/GitHub | Unit Testing',
        'focus_areas': 'Frontend Architecture | Backend Services | System Integration',
    },
    {
        'name': 'BlockBusters',
        'team_lead': 'Piotr Chebyrek',
        'department_name': 'Platform Engineering',
        'department_head': 'Sultan Suleyman',
        'github_link': 'https://github.com/sky/blockbusters',
        'status': PlatformTeam.TeamStatusChoices.REVIEW,
        'about_team': 'BlockBusters keeps shared service blocks ready for dependent teams.',
        'key_skills': 'Python | REST APIs | Docker | CI/CD',
        'focus_areas': 'Reliability | Automation | Release Management',
    },
]


def get_feedback_message(boundForm):
    if not boundForm.errors:
        return ''

    firstErrorList = next(iter(boundForm.errors.values()))
    return firstErrorList[0]


def is_admin_user(platformUser):
    return (
        platformUser.user_type == 'admin'
        or platformUser.is_staff
        or platformUser.is_superuser
    )


def get_profile_url(isNewMode=False):
    profileUrl = reverse('login_page:profile')

    if isNewMode:
        return f'{profileUrl}?mode=new'

    return profileUrl


def get_redirect_to(platformUser):
    if is_admin_user(platformUser):
        return reverse('login_page:admin_side')

    if not PlatformProfile.objects.filter(platform_user=platformUser).exists():
        return get_profile_url(isNewMode=True)

    return reverse('login_page:user_home')


def create_platform_user(signUpForm):
    user_model = get_user_model()
    signUpData = signUpForm.cleaned_data

    createdUser = user_model.objects.create_user(
        username=signUpData['username'],
        email=signUpData['email'],
        password=signUpData['password'],
        user_type='user',
    )

    return createdUser


def build_new_account_initial(platformUser):
    platformProfile = PlatformProfile.objects.filter(platform_user=platformUser).first()

    return {
        'full_name': platformProfile.full_name if platformProfile else '',
        'username': platformUser.username,
        'date_of_birth': platformProfile.date_of_birth if platformProfile else '',
        'gender': platformProfile.gender if platformProfile else '',
        'email': platformUser.email,
        'phone_number': platformProfile.phone_number if platformProfile else '',
        'status': platformProfile.status if platformProfile else '',
        'team_name': get_platform_profile_team_name(platformProfile),
        'team_role': platformProfile.team_role if platformProfile else '',
        'department_name': platformProfile.department_name if platformProfile else '',
        'department_head': platformProfile.department_head if platformProfile else '',
        'member_skills': platformProfile.member_skills if platformProfile else '',
    }


def save_platform_profile(platformUser, newAccountForm):
    newAccountData = newAccountForm.cleaned_data
    platformTeam = get_or_create_platform_team_for_profile_data(
        newAccountData['team_name'],
        fullName=newAccountData['full_name'],
        departmentName=newAccountData['department_name'],
        departmentHead=newAccountData['department_head'],
        memberSkills=newAccountData['member_skills'],
    )
    platformProfile, _ = PlatformProfile.objects.update_or_create(
        platform_user=platformUser,
        defaults={
            'platform_team': platformTeam,
            'full_name': newAccountData['full_name'],
            'date_of_birth': newAccountData['date_of_birth'],
            'gender': newAccountData['gender'],
            'phone_number': newAccountData['phone_number'],
            'status': newAccountData['status'],
            'team_name': newAccountData['team_name'],
            'team_role': newAccountData['team_role'],
            'department_name': newAccountData['department_name'],
            'department_head': newAccountData['department_head'],
            'member_skills': newAccountData['member_skills'],
        },
    )
    uploadedProfileImage = newAccountData.get('profile_image')
    if uploadedProfileImage:
        platformProfile.profile_image = uploadedProfileImage.read()
        platformProfile.profile_image_content_type = (uploadedProfileImage.content_type or '').strip().lower()
        platformProfile.save(update_fields=['profile_image', 'profile_image_content_type'])

    platformUser.username = newAccountData['username']
    platformUser.email = newAccountData['email']
    platformUser.save()

    # Keep account page data synchronized with profile edits when an account row already exists.
    platformAccount = PlatformAccount.objects.filter(platform_user=platformUser).first()
    if platformAccount:
        platformAccount.full_name = newAccountData['full_name']
        platformAccount.username = newAccountData['username']
        platformAccount.email = newAccountData['email']
        platformAccount.date_of_birth = newAccountData['date_of_birth']
        platformAccount.gender = newAccountData['gender']
        platformAccount.phone_number = newAccountData['phone_number']
        platformAccount.status = newAccountData['status']
        platformAccount.team_name = newAccountData['team_name']
        platformAccount.team_role = newAccountData['team_role']
        platformAccount.department_name = newAccountData['department_name']
        platformAccount.department_head = newAccountData['department_head']
        platformAccount.member_skills = newAccountData['member_skills']
        platformAccount.save(
            update_fields=[
                'full_name',
                'username',
                'email',
                'date_of_birth',
                'gender',
                'phone_number',
                'status',
                'team_name',
                'team_role',
                'department_name',
                'department_head',
                'member_skills',
            ]
        )

    return platformProfile


def build_account_initial(platformUser):
    platformAccount = PlatformAccount.objects.filter(platform_user=platformUser).first()
    platformProfile = PlatformProfile.objects.filter(platform_user=platformUser).first()

    return {
        'full_name': (
            platformAccount.full_name
            if platformAccount
            else platformProfile.full_name
            if platformProfile
            else ''
        ),
        'username': platformAccount.username if platformAccount else platformUser.username,
        'date_of_birth': (
            platformAccount.date_of_birth
            if platformAccount
            else platformProfile.date_of_birth
            if platformProfile
            else ''
        ),
        'gender': (
            platformAccount.gender
            if platformAccount
            else platformProfile.gender
            if platformProfile
            else ''
        ),
        'email': platformAccount.email if platformAccount else platformUser.email,
        'phone_number': (
            platformAccount.phone_number
            if platformAccount
            else platformProfile.phone_number
            if platformProfile
            else ''
        ),
        'status': (
            platformAccount.status
            if platformAccount
            else platformProfile.status
            if platformProfile
            else ''
        ),
        'team_name': (
            get_platform_profile_team_name(platformProfile)
            if platformProfile
            else platformAccount.team_name
            if platformAccount
            else ''
        ),
        'team_role': (
            platformAccount.team_role
            if platformAccount
            else platformProfile.team_role
            if platformProfile
            else ''
        ),
        'department_name': (
            platformAccount.department_name
            if platformAccount
            else platformProfile.department_name
            if platformProfile
            else ''
        ),
        'department_head': (
            platformAccount.department_head
            if platformAccount
            else platformProfile.department_head
            if platformProfile
            else ''
        ),
        'member_skills': (
            platformAccount.member_skills
            if platformAccount
            else platformProfile.member_skills
            if platformProfile
            else ''
        ),
    }


def save_platform_account(platformUser, accountForm):
    accountData = accountForm.cleaned_data
    platformTeam = get_or_create_platform_team_for_profile_data(
        accountData['team_name'],
        fullName=accountData['full_name'],
        departmentName=accountData['department_name'],
        departmentHead=accountData['department_head'],
        memberSkills=accountData['member_skills'],
    )
    profileDefaults = {
        'full_name': accountData['full_name'],
        'date_of_birth': accountData['date_of_birth'],
        'gender': accountData['gender'],
        'phone_number': accountData['phone_number'],
        'status': accountData['status'],
        'team_name': accountData['team_name'],
        'team_role': accountData['team_role'],
        'department_name': accountData['department_name'],
        'department_head': accountData['department_head'],
        'member_skills': accountData['member_skills'],
    }
    platformAccount, _ = PlatformAccount.objects.update_or_create(
        platform_user=platformUser,
        defaults={
            **profileDefaults,
            'username': accountData['username'],
            'email': accountData['email'],
        },
    )
    PlatformProfile.objects.update_or_create(
        platform_user=platformUser,
        defaults={
            **profileDefaults,
            'platform_team': platformTeam,
        },
    )

    platformUser.username = accountData['username']
    platformUser.email = accountData['email']
    platformUser.save()

    return platformAccount


def create_platform_user_from_account_form(accountForm):
    user_model = get_user_model()
    accountData = accountForm.cleaned_data
    createdUser = user_model.objects.create_user(
        username=accountData['username'],
        email=accountData['email'],
        password=accountData['password'],
        user_type='user',
    )
    save_platform_account(createdUser, accountForm)
    return createdUser


def reset_platform_password(resetPasswordForm):
    user_model = get_user_model()
    emailValue = resetPasswordForm.cleaned_data['email']
    matchedUser = user_model.objects.get(email__iexact=emailValue)
    oldPasswordHash = matchedUser.password
    matchedUser.set_password(resetPasswordForm.cleaned_data['new_password'])
    matchedUser.save()

    PlatformPasswordHistory.objects.create(
        platform_user=matchedUser,
        old_password_hash=oldPasswordHash,
        new_password_hash=matchedUser.password,
        change_source=PlatformPasswordHistory.ChangeSourceChoices.RESET,
    )

    return matchedUser


def get_platform_profile(platformUser):
    try:
        return platformUser.platform_profile
    except PlatformProfile.DoesNotExist:
        return None


def get_profile_image_data_url(platformProfile):
    if not platformProfile or not platformProfile.profile_image:
        return ''

    return get_binary_image_data_url(
        platformProfile.profile_image,
        platformProfile.profile_image_content_type,
    )


def get_binary_image_data_url(imageBytes, imageContentType):
    if not imageBytes:
        return ''

    normalizedContentType = (imageContentType or 'image/jpeg').strip().lower()
    encodedImage = base64.b64encode(bytes(imageBytes)).decode('ascii')
    return f'data:{normalizedContentType};base64,{encodedImage}'


def get_platform_team_image_data_url(platformTeam):
    if not platformTeam or not platformTeam.team_image:
        return ''

    return get_binary_image_data_url(
        platformTeam.team_image,
        platformTeam.team_image_content_type,
    )


def get_platform_profile_team_name(platformProfile):
    if not platformProfile:
        return ''

    platformTeam = getattr(platformProfile, 'platform_team', None)
    if platformTeam:
        return platformTeam.name

    return (platformProfile.team_name or '').strip()


def get_platform_team_by_name(teamName):
    normalizedTeamName = (teamName or '').strip()
    if not normalizedTeamName:
        return None

    return PlatformTeam.objects.filter(name__iexact=normalizedTeamName).first()


def build_platform_team_seed_from_profile_data(
    teamName,
    fullName='',
    departmentName='',
    departmentHead='',
    memberSkills='',
):
    normalizedTeamName = (teamName or '').strip()
    teamSlug = slugify(normalizedTeamName) or 'platform-team'
    normalizedFullName = (fullName or '').strip()
    normalizedDepartmentName = (departmentName or '').strip()
    normalizedDepartmentHead = (departmentHead or '').strip()
    normalizedMemberSkills = (memberSkills or '').strip()
    teamLead = normalizedFullName or 'Unassigned Team Lead'

    return {
        'name': normalizedTeamName,
        'team_lead': teamLead,
        'department_name': normalizedDepartmentName,
        'department_head': normalizedDepartmentHead,
        'github_link': f'https://github.com/sky/{teamSlug}',
        'status': PlatformTeam.TeamStatusChoices.ACTIVE,
        'git_project_name': normalizedTeamName,
        'jira_project_name': normalizedTeamName,
        'jira_board_link': f'https://jira.example.com/{teamSlug}',
        'wiki_link': f'https://wiki.example.com/{teamSlug}',
        'slack_link': 'https://slack.com',
        'daily_standup_time': '11:00',
        'about_team': (
            f'{normalizedTeamName} works inside {normalizedDepartmentName} '
            f'with {teamLead} as the first allocated contact.'
            if normalizedDepartmentName
            else f'{normalizedTeamName} has been seeded from user profile data.'
        ),
        'key_skills': normalizedMemberSkills,
        'focus_areas': 'Delivery Support | Team Coordination | System Integration',
    }


def fill_platform_team_blanks_from_profile_data(
    platformTeam,
    fullName='',
    departmentName='',
    departmentHead='',
    memberSkills='',
):
    fieldsToUpdate = []
    profileSeedValues = {
        'team_lead': (fullName or '').strip(),
        'department_name': (departmentName or '').strip(),
        'department_head': (departmentHead or '').strip(),
        'key_contact_1': (fullName or '').strip(),
        'key_contact_2': (departmentHead or '').strip(),
        'key_skills': (memberSkills or '').strip(),
    }

    for fieldName, profileSeedValue in profileSeedValues.items():
        if profileSeedValue and not (getattr(platformTeam, fieldName) or '').strip():
            setattr(platformTeam, fieldName, profileSeedValue)
            fieldsToUpdate.append(fieldName)

    if fieldsToUpdate:
        platformTeam.save(update_fields=[*fieldsToUpdate, 'updated_at'])

    return platformTeam


def get_or_create_platform_team_for_profile_data(
    teamName,
    fullName='',
    departmentName='',
    departmentHead='',
    memberSkills='',
):
    normalizedTeamName = (teamName or '').strip()
    if not normalizedTeamName:
        return None

    platformTeam = get_platform_team_by_name(normalizedTeamName)
    if platformTeam:
        return fill_platform_team_blanks_from_profile_data(
            platformTeam,
            fullName=fullName,
            departmentName=departmentName,
            departmentHead=departmentHead,
            memberSkills=memberSkills,
        )

    teamSeedItem = build_platform_team_seed_from_profile_data(
        normalizedTeamName,
        fullName=fullName,
        departmentName=departmentName,
        departmentHead=departmentHead,
        memberSkills=memberSkills,
    )
    return create_platform_team_from_seed(teamSeedItem)


def ensure_platform_team_for_profile(platformProfile):
    if not platformProfile:
        return None

    teamName = get_platform_profile_team_name(platformProfile)
    platformTeam = get_or_create_platform_team_for_profile_data(
        teamName,
        fullName=platformProfile.full_name,
        departmentName=platformProfile.department_name,
        departmentHead=platformProfile.department_head,
        memberSkills=platformProfile.member_skills,
    )
    if not platformTeam:
        return None

    if (
        platformProfile.platform_team_id != platformTeam.id
        or platformProfile.team_name != platformTeam.name
    ):
        PlatformProfile.objects.filter(pk=platformProfile.pk).update(
            platform_team=platformTeam,
            team_name=platformTeam.name,
        )
        platformProfile.platform_team = platformTeam
        platformProfile.team_name = platformTeam.name

    return platformTeam


def split_platform_team_entries(rawValue, limit=None):
    cleanedValue = (rawValue or '').strip()
    if not cleanedValue:
        return []

    teamEntries = [
        entryValue.strip()
        for entryValue in re.split(r'\s*(?:,|\||\n)\s*', cleanedValue)
        if entryValue.strip()
    ]

    if isinstance(limit, int) and limit > 0:
        return teamEntries[:limit]

    return teamEntries


def get_platform_team_unique_slug(teamName):
    baseSlug = slugify(teamName) or 'team'
    nextSlug = baseSlug
    nextIndex = 2

    while PlatformTeam.objects.filter(slug=nextSlug).exists():
        nextSlug = f'{baseSlug}-{nextIndex}'
        nextIndex += 1

    return nextSlug


def build_platform_team_seed_from_profiles():
    platformProfiles = (
        PlatformProfile.objects.select_related('platform_team')
        .filter(Q(platform_team__isnull=False) | ~Q(team_name=''))
        .order_by('platform_team__name', 'team_name', 'full_name')
    )
    teamSeedItems = []
    usedTeamNames = set()

    for platformProfile in platformProfiles:
        teamName = get_platform_profile_team_name(platformProfile)
        normalizedTeamName = teamName.lower()
        if not teamName or normalizedTeamName in usedTeamNames:
            continue

        usedTeamNames.add(normalizedTeamName)
        teamSeedItems.append(
            build_platform_team_seed_from_profile_data(
                teamName,
                fullName=platformProfile.full_name,
                departmentName=platformProfile.department_name,
                departmentHead=platformProfile.department_head,
                memberSkills=platformProfile.member_skills,
            )
        )

    return teamSeedItems


def normalize_platform_team_seed_time(seedValue):
    rawTime = (seedValue or '').strip()
    if not rawTime:
        return None

    try:
        return datetime.strptime(rawTime, '%H:%M').time()
    except ValueError:
        return None


def create_platform_team_from_seed(teamSeedItem):
    teamName = teamSeedItem['name']
    return PlatformTeam.objects.create(
        name=teamName,
        slug=get_platform_team_unique_slug(teamName),
        team_lead=teamSeedItem.get('team_lead', ''),
        department_name=teamSeedItem.get('department_name', ''),
        department_head=teamSeedItem.get('department_head', ''),
        github_link=teamSeedItem.get('github_link', ''),
        status=teamSeedItem.get('status') or PlatformTeam.TeamStatusChoices.ACTIVE,
        key_contact_1=teamSeedItem.get('team_lead', ''),
        key_contact_2=teamSeedItem.get('department_head', ''),
        jira_project_name=teamSeedItem.get('jira_project_name', ''),
        jira_board_link=teamSeedItem.get('jira_board_link', ''),
        git_project_name=teamSeedItem.get('git_project_name', ''),
        software_owned=teamSeedItem.get('software_owned', ''),
        versioning_approaches=teamSeedItem.get('versioning_approaches', ''),
        wiki_link=teamSeedItem.get('wiki_link', ''),
        wiki_search_terms=teamSeedItem.get('wiki_search_terms', ''),
        slack_channels=teamSeedItem.get('slack_channels', ''),
        slack_link=teamSeedItem.get('slack_link', ''),
        daily_standup_time=normalize_platform_team_seed_time(
            teamSeedItem.get('daily_standup_time', '')
        ),
        daily_standup_link=teamSeedItem.get('daily_standup_link', ''),
        about_team=teamSeedItem.get('about_team', ''),
        key_skills=teamSeedItem.get('key_skills', ''),
        focus_areas=teamSeedItem.get('focus_areas', ''),
    )


def ensure_platform_team_dependencies():
    platformTeams = list(PlatformTeam.objects.order_by('name'))
    if len(platformTeams) < 2:
        return

    for teamIndex, platformTeam in enumerate(platformTeams):
        if platformTeam.team_dependencies.exists():
            continue

        dependencyTeam = platformTeams[(teamIndex + 1) % len(platformTeams)]
        PlatformTeamDependency.objects.create(
            platform_team=platformTeam,
            dependency_name=dependencyTeam.name,
            is_upstream=(teamIndex % 2 == 0),
            status=dependencyTeam.status,
            updated_label='Recently',
        )


def sync_platform_profile_team_assignments():
    platformProfiles = (
        PlatformProfile.objects.select_related('platform_team')
        .filter(Q(platform_team__isnull=False) | ~Q(team_name=''))
        .order_by('team_name', 'full_name')
    )

    for platformProfile in platformProfiles:
        ensure_platform_team_for_profile(platformProfile)


def ensure_platform_teams():
    profileTeamSeed = build_platform_team_seed_from_profiles()
    teamSeedItems = profileTeamSeed or PLATFORM_TEAM_DEFAULT_SEED
    existingTeamNames = {
        teamName.lower()
        for teamName in PlatformTeam.objects.values_list('name', flat=True)
    }

    for teamSeedItem in teamSeedItems:
        normalizedTeamName = teamSeedItem['name'].strip().lower()
        if not normalizedTeamName or normalizedTeamName in existingTeamNames:
            continue

        create_platform_team_from_seed(teamSeedItem)
        existingTeamNames.add(normalizedTeamName)

    sync_platform_profile_team_assignments()
    ensure_platform_team_dependencies()


def get_platform_team_member_count_map():
    memberCountMap = {}
    platformProfiles = (
        PlatformProfile.objects.select_related('platform_team')
        .filter(Q(platform_team__isnull=False) | ~Q(team_name=''))
    )

    for platformProfile in platformProfiles:
        normalizedTeamName = get_platform_profile_team_name(platformProfile).lower()
        if not normalizedTeamName:
            continue

        memberCountMap[normalizedTeamName] = memberCountMap.get(normalizedTeamName, 0) + 1

    return memberCountMap


REPORTS_UNASSIGNED_MANAGER_VALUES = {
    'none',
    'no manager',
    'not assigned',
    'n/a',
    'na',
    'tbd',
    'unassigned',
    'unassigned team lead',
}


def get_reports_plural(countValue, singularLabel, pluralLabel=None):
    if countValue == 1:
        return f'{countValue} {singularLabel}'

    return f'{countValue} {pluralLabel or singularLabel + "s"}'


def is_platform_team_unmanaged(platformTeam):
    teamLead = (platformTeam.team_lead or '').strip()
    normalizedTeamLead = teamLead.lower()

    return (
        not teamLead
        or normalizedTeamLead in REPORTS_UNASSIGNED_MANAGER_VALUES
        or normalizedTeamLead.startswith('unassigned')
        or 'no manager' in normalizedTeamLead
    )


def get_reports_team_priority(memberCount, platformTeam):
    if memberCount >= 8 or platformTeam.status == PlatformTeam.TeamStatusChoices.REVIEW:
        return 'High'

    if memberCount >= 3 or platformTeam.status == PlatformTeam.TeamStatusChoices.ACTIVE:
        return 'Medium'

    return 'Low'


def get_reports_risk_level(priorityValue, platformTeam):
    if priorityValue == 'High':
        return 'High'

    if platformTeam.status == PlatformTeam.TeamStatusChoices.ACTIVE:
        return 'Medium'

    return 'Low'


def get_reports_acting_owner(platformTeam):
    ownerCandidates = [
        platformTeam.department_head,
        platformTeam.key_contact_1,
        platformTeam.key_contact_2,
    ]

    for ownerCandidate in ownerCandidates:
        cleanedOwner = (ownerCandidate or '').strip()
        if cleanedOwner:
            return cleanedOwner

    return 'None'


def build_reports_page_context(platformUser, generatedByName=''):
    ensure_platform_teams()

    if not generatedByName:
        platformProfile = get_platform_profile(platformUser)
        generatedByName = (
            (platformProfile.full_name if platformProfile else '').strip()
            or platformUser.get_full_name().strip()
            or platformUser.username
        )

    today = timezone.localdate()
    targetAssignmentDate = (today + timedelta(days=14)).strftime('%m/%d/%Y')
    reportingPeriod = today.strftime('%B %Y')
    platformTeams = list(PlatformTeam.objects.order_by('department_name', 'name'))
    memberCountMap = get_platform_team_member_count_map()
    totalTeams = len(platformTeams)
    activeTeams = sum(
        1
        for platformTeam in platformTeams
        if platformTeam.status == PlatformTeam.TeamStatusChoices.ACTIVE
    )
    reviewTeams = sum(
        1
        for platformTeam in platformTeams
        if platformTeam.status == PlatformTeam.TeamStatusChoices.REVIEW
    )
    inactiveTeams = sum(
        1
        for platformTeam in platformTeams
        if platformTeam.status == PlatformTeam.TeamStatusChoices.INACTIVE
    )
    totalMembers = sum(
        memberCountMap.get(platformTeam.name.strip().lower(), 0)
        for platformTeam in platformTeams
    )
    averageTeamSize = (totalMembers / totalTeams) if totalTeams else 0
    largestTeamSize = 0
    largestTeamName = 'No teams available'
    departmentStatsByName = {}

    for platformTeam in platformTeams:
        memberCount = memberCountMap.get(platformTeam.name.strip().lower(), 0)
        departmentName = (platformTeam.department_name or '').strip() or 'Unassigned Department'
        departmentStats = departmentStatsByName.setdefault(
            departmentName,
            {
                'name': departmentName,
                'teams': [],
                'total_teams': 0,
                'active_teams': 0,
                'review_teams': 0,
                'inactive_teams': 0,
                'member_total': 0,
                'largest_team_size': 0,
                'largest_team_name': 'None',
            },
        )

        departmentStats['teams'].append(platformTeam)
        departmentStats['total_teams'] += 1
        departmentStats['member_total'] += memberCount

        if platformTeam.status == PlatformTeam.TeamStatusChoices.ACTIVE:
            departmentStats['active_teams'] += 1
        elif platformTeam.status == PlatformTeam.TeamStatusChoices.REVIEW:
            departmentStats['review_teams'] += 1
        elif platformTeam.status == PlatformTeam.TeamStatusChoices.INACTIVE:
            departmentStats['inactive_teams'] += 1

        if memberCount >= departmentStats['largest_team_size']:
            departmentStats['largest_team_size'] = memberCount
            departmentStats['largest_team_name'] = platformTeam.name

        if memberCount >= largestTeamSize:
            largestTeamSize = memberCount
            largestTeamName = platformTeam.name

    departmentRows = []
    for departmentStats in sorted(
        departmentStatsByName.values(),
        key=lambda item: item['name'].lower(),
    ):
        totalDepartmentTeams = departmentStats['total_teams']
        averageDepartmentSize = (
            departmentStats['member_total'] / totalDepartmentTeams
            if totalDepartmentTeams
            else 0
        )
        noteParts = []
        if departmentStats['review_teams']:
            noteParts.append(get_reports_plural(departmentStats['review_teams'], 'team') + ' on review')
        if departmentStats['inactive_teams']:
            noteParts.append(get_reports_plural(departmentStats['inactive_teams'], 'inactive team'))

        departmentRows.append(
            {
                'name': departmentStats['name'],
                'total_teams': totalDepartmentTeams,
                'active_teams': departmentStats['active_teams'],
                'inactive_teams': departmentStats['inactive_teams'],
                'average_team_size': f'{averageDepartmentSize:.1f}',
                'largest_team_size': departmentStats['largest_team_size'],
                'largest_team_name': departmentStats['largest_team_name'],
                'notes': ', '.join(noteParts) if noteParts else 'None',
            }
        )

    topDepartment = max(
        departmentRows,
        key=lambda item: (item['total_teams'], item['name']),
        default={'name': 'No department', 'total_teams': 0},
    )

    unmanagedRows = []
    for platformTeam in platformTeams:
        if not is_platform_team_unmanaged(platformTeam):
            continue

        memberCount = memberCountMap.get(platformTeam.name.strip().lower(), 0)
        priority = get_reports_team_priority(memberCount, platformTeam)
        riskLevel = get_reports_risk_level(priority, platformTeam)

        unmanagedRows.append(
            {
                'name': platformTeam.name,
                'department_name': (
                    (platformTeam.department_name or '').strip()
                    or 'Unassigned Department'
                ),
                'member_count': memberCount,
                'status_display': platformTeam.get_status_display(),
                'priority': priority,
                'manager_name': 'None',
                'acting_owner': get_reports_acting_owner(platformTeam),
                'target_assignment_date': targetAssignmentDate,
                'risk_level': riskLevel,
                'notes': 'No manager assigned',
            }
        )

    unmanagedCount = len(unmanagedRows)
    unmanagedDepartments = {
        unmanagedRow['department_name']
        for unmanagedRow in unmanagedRows
    }
    unmanagedDepartmentCount = len(unmanagedDepartments)
    unmanagedPercentage = (unmanagedCount / totalTeams * 100) if totalTeams else 0
    highPriorityCases = sum(
        1
        for unmanagedRow in unmanagedRows
        if unmanagedRow['priority'] == 'High'
    )
    managerCoverageCount = totalTeams - unmanagedCount
    managerCoveragePercentage = (managerCoverageCount / totalTeams * 100) if totalTeams else 0

    stats = {
        'total_teams': totalTeams,
        'active_teams': activeTeams,
        'review_teams': reviewTeams,
        'inactive_teams': inactiveTeams,
        'departments_covered': len(departmentRows),
        'average_team_size': f'{averageTeamSize:.1f}',
        'largest_team_size': largestTeamSize,
        'largest_team_name': largestTeamName,
        'top_department': topDepartment['name'],
        'top_department_team_count': topDepartment['total_teams'],
        'reporting_period': reportingPeriod,
        'total_members': totalMembers,
        'unmanaged_teams': unmanagedCount,
        'unmanaged_departments': unmanagedDepartmentCount,
        'unmanaged_percentage': f'{unmanagedPercentage:.1f}%',
        'high_priority_cases': highPriorityCases,
        'manager_coverage_percentage': f'{managerCoveragePercentage:.1f}%',
    }

    if unmanagedCount:
        unmanagedBody = (
            f'At present, {get_reports_plural(unmanagedCount, "team")} '
            f'are operating without a designated manager across '
            f'{get_reports_plural(unmanagedDepartmentCount, "department")}. '
            f'Priority assignment is recommended for '
            f'{get_reports_plural(highPriorityCases, "high-priority case")} '
            'due to delivery dependency and accountability risk.'
        )
        unmanagedImmediateAction = (
            'Immediate Action Recommendation: Assign acting owners for all '
            'high-priority unmanaged teams within 5 business days.'
        )
    else:
        unmanagedBody = (
            f'At present, all {get_reports_plural(totalTeams, "tracked team")} '
            'have assigned managers. Continue reviewing ownership weekly to '
            'keep escalation paths and delivery accountability clear.'
        )
        unmanagedImmediateAction = (
            'Immediate Action Recommendation: No manager coverage gaps are '
            'currently recorded. Continue monitoring team ownership weekly.'
        )

    teamCountBody = (
        f'As of the selected reporting period, the organisation maintains '
        f'{get_reports_plural(totalTeams, "team")} across '
        f'{get_reports_plural(len(departmentRows), "department")}. Of these, '
        f'{activeTeams} are active, {reviewTeams} are on review, and '
        f'{inactiveTeams} are inactive. The average team size is '
        f'{averageTeamSize:.1f} members, with the largest team being '
        f'{largestTeamName} at {get_reports_plural(largestTeamSize, "member")}. '
        f'Team distribution is currently led by {topDepartment["name"]}.'
    )
    summaryBody = (
        f'Within the selected period, {get_reports_plural(activeTeams, "active team")} '
        f'are tracked across {get_reports_plural(len(departmentRows), "department")}. '
        f'The platform records {get_reports_plural(totalMembers, "profiled member")} '
        f'and manager coverage is {managerCoveragePercentage:.1f}%. Major focus '
        'items for the next cycle include manager coverage, department balance, '
        'and release readiness.'
    )
    unmanagedManagerVerb = 'has' if unmanagedCount == 1 else 'have'
    unmanagedReviewVerb = 'requires' if unmanagedCount == 1 else 'require'

    return {
        'reports_generated_by': generatedByName,
        'reports_stats': stats,
        'reports_department_rows': departmentRows,
        'reports_unmanaged_rows': unmanagedRows,
        'reports_has_unmanaged_rows': bool(unmanagedRows),
        'reports_team_count_body': teamCountBody,
        'reports_summary_body': summaryBody,
        'reports_unmanaged_body': unmanagedBody,
        'reports_summary_executive': (
            f'This summary report provides a consolidated overview of '
            f'{get_reports_plural(totalTeams, "team")}, '
            f'{get_reports_plural(len(departmentRows), "department")}, and '
            f'{get_reports_plural(totalMembers, "profiled member")} for '
            f'{reportingPeriod}. It is intended for leadership review and '
            'highlights areas that may require action to improve accountability '
            'and team governance.'
        ),
        'reports_summary_key_insights': (
            f'- {activeTeams} teams are active, {reviewTeams} are on review, '
            f'and {inactiveTeams} are inactive.\n'
            f'- {get_reports_plural(unmanagedCount, "team")} currently '
            f'{unmanagedManagerVerb} no assigned manager.\n'
            f'- {topDepartment["name"]} has the highest team count.\n'
            f'- Average team size is {averageTeamSize:.1f} members.'
        ),
        'reports_summary_risks': (
            f'- Manager coverage is currently {managerCoveragePercentage:.1f}%.\n'
            '- Unassigned ownership can delay approvals and escalation.\n'
            '- Larger unmanaged teams present higher coordination risk.\n'
            '- Teams on review may require closer leadership follow-up.'
        ),
        'reports_summary_recommended_actions': (
            '1. Assign a manager or acting lead to all active unmanaged teams.\n'
            '2. Prioritize high-headcount or customer-facing teams for immediate ownership assignment.\n'
            '3. Keep every active team tied to a named accountable owner.\n'
            '4. Review unresolved ownership gaps weekly until closed.'
        ),
        'reports_unmanaged_immediate_action': unmanagedImmediateAction,
        'reports_unmanaged_observations': (
            f'- {get_reports_plural(unmanagedCount, "team")} currently '
            f'{unmanagedReviewVerb} ownership review.\n'
            f'- {get_reports_plural(highPriorityCases, "high-priority case")} '
            'should be handled first.\n'
            '- Track unmanaged status weekly until resolved.'
        ),
        'reports_unmanaged_required_actions': (
            '1. Assign acting manager or team owner within 5 business days.\n'
            '2. Finalize permanent manager assignment for high-priority teams within 2 weeks.\n'
            '3. Confirm action owner and target assignment date for each listed team.\n'
            '4. Escalate unresolved ownership gaps to department leadership.'
        ),
    }


def get_platform_team_status_class(statusValue):
    normalizedStatus = (statusValue or '').strip().lower()
    if normalizedStatus in {'active', 'review', 'inactive'}:
        return normalizedStatus

    return 'active'


def serialize_platform_team_card(platformTeam, memberCountMap):
    memberCount = memberCountMap.get(platformTeam.name.strip().lower(), 0)

    return {
        'id': platformTeam.id,
        'name': platformTeam.name,
        'team_lead': platformTeam.team_lead,
        'department_name': platformTeam.department_name,
        'member_count': memberCount,
        'member_count_label': (
            f'{memberCount} Member'
            if memberCount == 1
            else f'{memberCount} Members'
        ),
        'github_link': platformTeam.github_link,
        'status': platformTeam.status,
        'status_display': platformTeam.get_status_display(),
        'status_class': get_platform_team_status_class(platformTeam.status),
        'detail_url': reverse(
            'login_page:user_home_tool_team_detail',
            kwargs={'team_slug': platformTeam.slug},
        ),
    }


def search_user_home_team_directory(queryValue, limit=8):
    cleanedQuery = (queryValue or '').strip()
    if len(cleanedQuery) < 3:
        return []

    ensure_platform_teams()

    matchingTeams = list(
        PlatformTeam.objects.filter(
            Q(name__icontains=cleanedQuery)
            | Q(department_name__icontains=cleanedQuery)
            | Q(team_lead__icontains=cleanedQuery)
            | Q(department_head__icontains=cleanedQuery)
        )
        .order_by('name')
        .distinct()[:limit]
    )

    searchResults = []
    usedDepartmentNames = set()

    for platformTeam in matchingTeams:
        departmentName = (platformTeam.department_name or '').strip()
        if departmentName and cleanedQuery.lower() in departmentName.lower():
            normalizedDepartmentName = departmentName.lower()
            if normalizedDepartmentName not in usedDepartmentNames:
                usedDepartmentNames.add(normalizedDepartmentName)
                departmentTeamCount = PlatformTeam.objects.filter(
                    department_name__iexact=departmentName
                ).count()
                searchResults.append(
                    {
                        'result_type': 'department',
                        'full_name': departmentName,
                        'team_role': 'Department',
                        'status': (
                            f'{departmentTeamCount} Team'
                            if departmentTeamCount == 1
                            else f'{departmentTeamCount} Teams'
                        ),
                    }
                )

        searchResults.append(
            {
                'result_type': 'team',
                'full_name': platformTeam.name,
                'team_role': departmentName or platformTeam.team_lead or 'Team',
                'status': platformTeam.get_status_display(),
                'detail_url': reverse(
                    'login_page:user_home_tool_team_detail',
                    kwargs={'team_slug': platformTeam.slug},
                ),
            }
        )

        if len(searchResults) >= limit:
            break

    return searchResults[:limit]


def build_teams_page_context(platformUser):
    ensure_platform_teams()
    platformTeams = list(PlatformTeam.objects.order_by('name'))
    memberCountMap = get_platform_team_member_count_map()
    teamsPageContext = build_user_home_context(platformUser)
    teamsPageContext.update(
        {
            'teams': [
                serialize_platform_team_card(platformTeam, memberCountMap)
                for platformTeam in platformTeams
            ],
            'dashboard_url': reverse('login_page:user_home'),
            'user_home_search_url': reverse('login_page:user_home_team_directory_search'),
            'user_home_search_config': {
                'resultsTitle': 'Teams and departments',
                'emptyMessage': 'No matching teams or departments found yet.',
            },
        }
    )

    return teamsPageContext


def get_platform_team_initials(teamName):
    teamWords = [part for part in re.split(r'\s+', (teamName or '').strip()) if part]
    if not teamWords:
        return 'NA'
    if len(teamWords) == 1:
        return teamWords[0][:2].upper()

    return f'{teamWords[0][0]}{teamWords[1][0]}'.upper()


def build_platform_team_dependency_outgoing_map(platformTeams):
    outgoingMap = {
        platformTeam.name: []
        for platformTeam in platformTeams
    }
    teamNamesByKey = {
        platformTeam.name.strip().lower(): platformTeam.name
        for platformTeam in platformTeams
    }
    dependencyRows = (
        PlatformTeamDependency.objects.filter(platform_team__in=platformTeams)
        .select_related('platform_team')
        .order_by('platform_team__name', 'dependency_name')
    )

    for dependencyRow in dependencyRows:
        sourceTeamName = dependencyRow.platform_team.name
        dependencyName = (dependencyRow.dependency_name or '').strip()
        if not dependencyName:
            continue

        dependencyKey = dependencyName.lower()
        if dependencyRow.is_upstream and dependencyKey in teamNamesByKey:
            mappedSourceTeamName = teamNamesByKey[dependencyKey]
            outgoingMap.setdefault(mappedSourceTeamName, []).append(sourceTeamName)
            continue

        outgoingMap.setdefault(sourceTeamName, []).append(dependencyName)

    for sourceTeamName, dependencyNames in outgoingMap.items():
        uniqueDependencyNames = []
        usedDependencyKeys = set()
        for dependencyName in dependencyNames:
            dependencyKey = dependencyName.strip().lower()
            if not dependencyKey or dependencyKey in usedDependencyKeys:
                continue

            usedDependencyKeys.add(dependencyKey)
            uniqueDependencyNames.append(dependencyName)

        outgoingMap[sourceTeamName] = uniqueDependencyNames

    return outgoingMap


def serialize_platform_organisation_row(platformTeam, outgoingDependencyMap):
    jiraProjectName = (
        (platformTeam.jira_project_name or '').strip()
        or (platformTeam.git_project_name or '').strip()
        or slugify(platformTeam.name).replace('-', '').upper()
    )

    return {
        'Department': (platformTeam.department_name or '').strip() or 'Unassigned Department',
        'Department Head': (platformTeam.department_head or '').strip() or 'Department Head',
        'Team Name': platformTeam.name,
        'Team Leader': (platformTeam.team_lead or '').strip() or 'Team Leader',
        'Jira Project Name': jiraProjectName,
        'Downstream Dependencies': outgoingDependencyMap.get(platformTeam.name, []),
        'Avatar': get_platform_team_initials(platformTeam.name),
    }


def build_organisation_page_context(platformUser):
    ensure_platform_teams()
    platformTeams = list(PlatformTeam.objects.order_by('name'))
    outgoingDependencyMap = build_platform_team_dependency_outgoing_map(platformTeams)
    organisationPageContext = build_user_home_context(platformUser)
    organisationPageContext.update(
        {
            'dashboard_url': reverse('login_page:user_home'),
            'organisation_team_rows': [
                serialize_platform_organisation_row(platformTeam, outgoingDependencyMap)
                for platformTeam in platformTeams
            ],
        }
    )

    return organisationPageContext


def serialize_platform_data_visualisation_record(platformTeam, profileSkillsByTeam):
    jiraProjectName = (
        (platformTeam.jira_project_name or '').strip()
        or (platformTeam.git_project_name or '').strip()
        or slugify(platformTeam.name).replace('-', '').upper()
        or 'Unassigned'
    )
    normalizedTeamName = platformTeam.name.strip().lower()
    skillEntries = [
        (platformTeam.key_skills or '').strip(),
        *profileSkillsByTeam.get(normalizedTeamName, []),
    ]
    keySkills = ' | '.join(skillEntry for skillEntry in skillEntries if skillEntry)

    return {
        'Department': (platformTeam.department_name or '').strip() or 'Unassigned Department',
        'Jira Project Name': jiraProjectName,
        'Key Skills & Technologies': keySkills or 'Unspecified',
        'Department Head': (platformTeam.department_head or '').strip() or 'Department Head',
        'Team Name': platformTeam.name,
    }


def build_platform_profile_skills_by_team():
    profileSkillsByTeam = {}
    platformProfiles = (
        PlatformProfile.objects.select_related('platform_team')
        .filter(Q(platform_team__isnull=False) | ~Q(team_name=''))
        .order_by('platform_team__name', 'team_name', 'full_name')
    )

    for platformProfile in platformProfiles:
        normalizedTeamName = get_platform_profile_team_name(platformProfile).lower()
        memberSkills = (platformProfile.member_skills or '').strip()
        if not normalizedTeamName or not memberSkills:
            continue

        profileSkillsByTeam.setdefault(normalizedTeamName, []).append(memberSkills)

    return profileSkillsByTeam


def build_data_visualisation_page_context(platformUser):
    ensure_platform_teams()
    platformTeams = list(PlatformTeam.objects.order_by('department_name', 'name'))
    profileSkillsByTeam = build_platform_profile_skills_by_team()
    dataVisualisationPageContext = build_user_home_context(platformUser)
    dataVisualisationPageContext.update(
        {
            'dashboard_url': reverse('login_page:user_home'),
            'data_visualisation_records': [
                serialize_platform_data_visualisation_record(
                    platformTeam,
                    profileSkillsByTeam,
                )
                for platformTeam in platformTeams
            ],
        }
    )

    return dataVisualisationPageContext


def get_user_home_compose_url(emailValue=''):
    inboxUrl = reverse('login_page:user_home_tool_message')
    normalizedEmail = (emailValue or '').strip()
    if not normalizedEmail:
        return inboxUrl

    return f'{inboxUrl}?{urlencode({"compose_to": normalizedEmail})}'


def get_platform_contact_email(contactValue):
    normalizedContact = (contactValue or '').strip()
    if not normalizedContact:
        return ''

    if EMAIL_ADDRESS_PATTERN.match(normalizedContact):
        return normalizedContact.lower()

    matchedProfile = (
        PlatformProfile.objects.select_related('platform_user')
        .filter(full_name__iexact=normalizedContact)
        .first()
    )
    if matchedProfile and matchedProfile.platform_user.email:
        return matchedProfile.platform_user.email

    user_model = get_user_model()
    matchedUser = (
        user_model.objects.filter(
            Q(email__iexact=normalizedContact) | Q(username__iexact=normalizedContact)
        )
        .first()
    )
    if matchedUser and matchedUser.email:
        return matchedUser.email

    return ''


def serialize_platform_team_contact(contactValue, fallbackName=''):
    contactName = (contactValue or fallbackName or '').strip()
    contactEmail = get_platform_contact_email(contactName)

    return {
        'name': contactName,
        'email': contactEmail,
        'message_url': get_user_home_compose_url(contactEmail),
    }


def serialize_platform_team_member(platformProfile):
    memberEmail = platformProfile.platform_user.email

    return {
        'full_name': platformProfile.full_name,
        'team_role': platformProfile.team_role,
        'status': platformProfile.status,
        'email': memberEmail,
        'message_url': get_user_home_compose_url(memberEmail),
        'member_skills': platformProfile.member_skills,
        'avatar_url': get_profile_image_data_url(platformProfile),
    }


def serialize_platform_team_dependency(platformTeamDependency):
    directionValue = 'upstream' if platformTeamDependency.is_upstream else 'downstream'

    return {
        'name': platformTeamDependency.dependency_name,
        'direction': directionValue,
        'direction_display': 'Upstream' if platformTeamDependency.is_upstream else 'Downstream',
        'status': platformTeamDependency.status,
        'status_display': platformTeamDependency.get_status_display(),
        'status_class': get_platform_team_status_class(platformTeamDependency.status),
        'updated_label': platformTeamDependency.updated_label,
        'is_upstream': platformTeamDependency.is_upstream,
    }


def build_team_page_context(platformUser, platformTeam):
    ensure_platform_teams()
    platformTeam.refresh_from_db()
    teamMembers = [
        serialize_platform_team_member(platformProfile)
        for platformProfile in (
            PlatformProfile.objects.filter(
                Q(platform_team=platformTeam)
                | Q(team_name__iexact=platformTeam.name)
            )
            .select_related('platform_user', 'platform_team')
            .distinct()
            .order_by('full_name')
        )
    ]
    teamDependencies = [
        serialize_platform_team_dependency(platformTeamDependency)
        for platformTeamDependency in platformTeam.team_dependencies.all()
    ]
    firstDependency = teamDependencies[0] if teamDependencies else None
    departmentHeadContact = serialize_platform_team_contact(
        platformTeam.department_head,
        'Department Head name',
    )
    keyContactOne = serialize_platform_team_contact(
        platformTeam.key_contact_1,
        platformTeam.team_lead,
    )
    keyContactTwo = serialize_platform_team_contact(
        platformTeam.key_contact_2,
        platformTeam.department_head,
    )
    teamPageContext = build_user_home_context(platformUser)
    teamPageContext.update(
        {
            'platform_team': platformTeam,
            'team_image_url': get_platform_team_image_data_url(platformTeam),
            'dashboard_url': reverse('login_page:user_home'),
            'teams_url': reverse('login_page:user_home_tool_team'),
            'team_settings_action_url': reverse(
                'login_page:user_home_tool_team_detail',
                kwargs={'team_slug': platformTeam.slug},
            ),
            'team_members': teamMembers,
            'team_member_count': len(teamMembers),
            'team_member_count_label': (
                f'{len(teamMembers)} Member'
                if len(teamMembers) == 1
                else f'{len(teamMembers)} Members'
            ),
            'team_department_head_contact': departmentHeadContact,
            'team_key_contact_1': keyContactOne,
            'team_key_contact_2': keyContactTwo,
            'team_dependencies': teamDependencies,
            'team_first_dependency': firstDependency,
            'team_dependency_settings_name': firstDependency['name'] if firstDependency else '',
            'team_dependency_settings_type': firstDependency['direction_display'] if firstDependency else '',
            'team_key_skills': split_platform_team_entries(platformTeam.key_skills),
            'team_focus_areas': split_platform_team_entries(platformTeam.focus_areas),
            'team_standup_time_value': (
                platformTeam.daily_standup_time.strftime('%H:%M')
                if platformTeam.daily_standup_time
                else ''
            ),
        }
    )

    return teamPageContext


def save_platform_team_settings(platformTeam, teamSettingsForm):
    teamSettingsData = teamSettingsForm.cleaned_data
    teamFieldsToUpdate = []
    teamFieldNames = [
        'key_contact_1',
        'key_contact_2',
        'jira_project_name',
        'jira_board_link',
        'git_project_name',
        'github_link',
        'software_owned',
        'versioning_approaches',
        'wiki_link',
        'wiki_search_terms',
        'slack_channels',
        'slack_link',
        'daily_standup_time',
        'daily_standup_link',
        'about_team',
        'key_skills',
        'focus_areas',
    ]

    for fieldName in teamFieldNames:
        nextValue = teamSettingsData.get(fieldName)
        if isinstance(nextValue, str):
            nextValue = nextValue.strip()

        if getattr(platformTeam, fieldName) == nextValue:
            continue

        setattr(platformTeam, fieldName, nextValue)
        teamFieldsToUpdate.append(fieldName)

    if teamFieldsToUpdate:
        platformTeam.save(update_fields=[*teamFieldsToUpdate, 'updated_at'])

    uploadedTeamImage = teamSettingsData.get('team_image')
    if uploadedTeamImage:
        platformTeam.team_image = uploadedTeamImage.read()
        platformTeam.team_image_content_type = (
            uploadedTeamImage.content_type or ''
        ).strip().lower()
        platformTeam.save(
            update_fields=['team_image', 'team_image_content_type', 'updated_at']
        )

    dependencyName = (teamSettingsData.get('dependency_name') or '').strip()
    dependencyType = teamSettingsData.get('dependency_type') or ''
    if dependencyName:
        isUpstream = dependencyType != 'downstream'
        platformTeamDependency = platformTeam.team_dependencies.order_by('id').first()
        if platformTeamDependency:
            dependencyFieldsToUpdate = []
            if platformTeamDependency.dependency_name != dependencyName:
                platformTeamDependency.dependency_name = dependencyName
                dependencyFieldsToUpdate.append('dependency_name')
            if platformTeamDependency.is_upstream != isUpstream:
                platformTeamDependency.is_upstream = isUpstream
                dependencyFieldsToUpdate.append('is_upstream')

            if dependencyFieldsToUpdate:
                platformTeamDependency.save(update_fields=dependencyFieldsToUpdate)
        else:
            PlatformTeamDependency.objects.create(
                platform_team=platformTeam,
                dependency_name=dependencyName,
                is_upstream=isUpstream,
                status=platformTeam.status,
                updated_label='Recently',
            )

    return platformTeam


def build_user_home_inbox_seed(platformUser):
    platformProfile = get_platform_profile(platformUser)
    profileName = platformProfile.full_name if platformProfile else platformUser.username
    teamName = get_platform_profile_team_name(platformProfile) or 'Your team'

    return [
        {
            'message_mode': PlatformInboxMessage.MessageModeChoices.INBOX,
            'draft_type': '',
            'sender_name': 'IT Support',
            'sender_email': 'it.support@sky.com',
            'recipient_name': profileName,
            'recipient_email': platformUser.email,
            'email_subject': 'E-mail reset password',
            'email_body': 'Please reset your password before Friday. Use the recovery portal and confirm once done.',
            'previous_message': 'Please reset your password before Friday. Use the recovery portal and confirm once done.',
            'email_reply': '',
            'is_reply': False,
            'is_read': False,
        },
        {
            'message_mode': PlatformInboxMessage.MessageModeChoices.INBOX,
            'draft_type': '',
            'sender_name': 'Project Coordinator',
            'sender_email': 'project.coordinator@sky.com',
            'recipient_name': profileName,
            'recipient_email': platformUser.email,
            'email_subject': 'Team project deadline',
            'email_body': 'Reminder that your team deliverable is due at 17:00 on Friday. Reply if you need clarification.',
            'previous_message': 'Reminder that your team deliverable is due at 17:00 on Friday. Reply if you need clarification.',
            'email_reply': '',
            'is_reply': False,
            'is_read': False,
        },
        {
            'message_mode': PlatformInboxMessage.MessageModeChoices.INBOX,
            'draft_type': '',
            'sender_name': 'Scrum Lead',
            'sender_email': 'scrum.lead@sky.com',
            'recipient_name': profileName,
            'recipient_email': platformUser.email,
            'email_subject': 'Project status update',
            'email_body': 'Share your current sprint progress with blockers and expected completion date.',
            'previous_message': 'Share your current sprint progress with blockers and expected completion date.',
            'email_reply': '',
            'is_reply': False,
            'is_read': True,
        },
        {
            'message_mode': PlatformInboxMessage.MessageModeChoices.INBOX,
            'draft_type': '',
            'sender_name': 'HR Team',
            'sender_email': 'hr.team@sky.com',
            'recipient_name': profileName,
            'recipient_email': platformUser.email,
            'email_subject': 'New team member introduction',
            'email_body': 'A new member joins next week. Please welcome them and share onboarding notes.',
            'previous_message': 'A new member joins next week. Please welcome them and share onboarding notes.',
            'email_reply': '',
            'is_reply': False,
            'is_read': True,
        },
        {
            'message_mode': PlatformInboxMessage.MessageModeChoices.INBOX,
            'draft_type': '',
            'sender_name': 'System Administration',
            'sender_email': 'system.administration@sky.com',
            'recipient_name': profileName,
            'recipient_email': platformUser.email,
            'email_subject': 'System maintenance notice',
            'email_body': 'Planned maintenance will happen this weekend from 22:00 to 02:00. Some services may be unavailable.',
            'previous_message': 'Planned maintenance will happen this weekend from 22:00 to 02:00. Some services may be unavailable.',
            'email_reply': '',
            'is_reply': False,
            'is_read': True,
        },
        {
            'message_mode': PlatformInboxMessage.MessageModeChoices.SENT,
            'draft_type': PlatformInboxMessage.DraftTypeChoices.COMPOSE,
            'sender_name': profileName,
            'sender_email': platformUser.email,
            'recipient_name': 'Operations Team',
            'recipient_email': 'operations.team@sky.com',
            'email_subject': 'Weekly progress update',
            'email_body': f'The {teamName} team has completed sprint tasks and started integration testing.',
            'previous_message': '',
            'email_reply': '',
            'is_reply': False,
            'is_read': True,
        },
        {
            'message_mode': PlatformInboxMessage.MessageModeChoices.SENT,
            'draft_type': PlatformInboxMessage.DraftTypeChoices.REPLY,
            'sender_name': profileName,
            'sender_email': platformUser.email,
            'recipient_name': 'IT Support',
            'recipient_email': 'it.support@sky.com',
            'email_subject': 'Access request',
            'email_body': 'Access request details have been submitted in the portal.',
            'previous_message': 'Please provide details for your access request.',
            'email_reply': '',
            'is_reply': True,
            'is_read': True,
        },
        {
            'message_mode': PlatformInboxMessage.MessageModeChoices.DRAFTS,
            'draft_type': PlatformInboxMessage.DraftTypeChoices.COMPOSE,
            'sender_name': profileName,
            'sender_email': platformUser.email,
            'recipient_name': 'Project Team',
            'recipient_email': 'project.team@sky.com',
            'email_subject': 'Draft: sprint priorities',
            'email_body': 'Collecting sprint priorities for next planning session.',
            'previous_message': '',
            'email_reply': '',
            'is_reply': False,
            'is_read': True,
        },
        {
            'message_mode': PlatformInboxMessage.MessageModeChoices.DRAFTS,
            'draft_type': PlatformInboxMessage.DraftTypeChoices.REPLY,
            'sender_name': profileName,
            'sender_email': platformUser.email,
            'recipient_name': 'System Administration',
            'recipient_email': 'system.administration@sky.com',
            'email_subject': 'Draft reply: maintenance notice',
            'email_body': 'Thanks for the notice, can you confirm service availability windows?',
            'previous_message': 'Planned maintenance will happen this weekend from 22:00 to 02:00. Some services may be unavailable.',
            'email_reply': '',
            'is_reply': True,
            'is_read': True,
        },
    ]


def ensure_user_home_inbox_messages(platformUser):
    existingModes = set(
        PlatformInboxMessage.objects.filter(platform_user=platformUser)
        .values_list('message_mode', flat=True)
    )

    for inboxSeedItem in build_user_home_inbox_seed(platformUser):
        if inboxSeedItem['message_mode'] in existingModes:
            continue

        PlatformInboxMessage.objects.create(
            platform_user=platformUser,
            message_mode=inboxSeedItem['message_mode'],
            draft_type=inboxSeedItem['draft_type'],
            sender_name=inboxSeedItem['sender_name'],
            sender_email=inboxSeedItem['sender_email'],
            recipient_name=inboxSeedItem['recipient_name'],
            recipient_email=inboxSeedItem['recipient_email'],
            email_subject=inboxSeedItem['email_subject'],
            email_body=inboxSeedItem['email_body'],
            previous_message=inboxSeedItem['previous_message'],
            email_reply=inboxSeedItem['email_reply'],
            is_reply=inboxSeedItem['is_reply'],
            is_read=inboxSeedItem['is_read'],
            replied_at=timezone.now() if inboxSeedItem['is_reply'] else None,
        )


def get_user_home_inbox_messages(platformUser, limit=None):
    ensure_user_home_inbox_messages(platformUser)
    inboxQuerySet = PlatformInboxMessage.objects.filter(
        platform_user=platformUser,
        message_mode=PlatformInboxMessage.MessageModeChoices.INBOX,
        is_hidden_from_user=False,
    ).order_by('-created_at', '-id')
    if limit is not None:
        inboxQuerySet = inboxQuerySet[:limit]

    return list(inboxQuerySet)


def get_user_home_selected_inbox_message(inboxMessages, selectedMessageId):
    if not inboxMessages:
        return None

    try:
        normalizedSelectedMessageId = int((str(selectedMessageId or '')).strip())
    except (TypeError, ValueError):
        normalizedSelectedMessageId = None

    if normalizedSelectedMessageId is None:
        return inboxMessages[0]

    for inboxMessage in inboxMessages:
        if inboxMessage.id == normalizedSelectedMessageId:
            return inboxMessage

    return inboxMessages[0]


def get_user_home_notification_messages(platformUser, limit=USER_HOME_NOTIFICATION_VISIBLE_COUNT):
    ensure_user_home_inbox_messages(platformUser)
    return list(
        PlatformInboxMessage.objects.filter(
            platform_user=platformUser,
            message_mode=PlatformInboxMessage.MessageModeChoices.INBOX,
            is_hidden_from_user=False,
        )
        .order_by('is_read', '-created_at', '-id')[:limit]
    )


def get_user_home_unread_message_count(platformUser):
    ensure_user_home_inbox_messages(platformUser)
    return PlatformInboxMessage.objects.filter(
        platform_user=platformUser,
        message_mode=PlatformInboxMessage.MessageModeChoices.INBOX,
        is_read=False,
        is_hidden_from_user=False,
    ).count()


def mark_user_home_inbox_messages_as_read(platformUser):
    ensure_user_home_inbox_messages(platformUser)
    return PlatformInboxMessage.objects.filter(
        platform_user=platformUser,
        message_mode=PlatformInboxMessage.MessageModeChoices.INBOX,
        is_read=False,
        is_hidden_from_user=False,
    ).update(is_read=True)


def save_user_home_inbox_reply(platformUser, messageId, emailReplyValue):
    try:
        normalizedMessageId = int((str(messageId or '')).strip())
    except (TypeError, ValueError):
        return None

    inboxMessage = PlatformInboxMessage.objects.filter(
        platform_user=platformUser,
        pk=normalizedMessageId,
    ).first()
    if inboxMessage is None:
        return None

    normalizedReplyValue = (emailReplyValue or '').strip()
    inboxMessage.email_reply = normalizedReplyValue
    inboxMessage.is_reply = bool(normalizedReplyValue)
    inboxMessage.is_read = True
    inboxMessage.replied_at = timezone.now() if normalizedReplyValue else None
    inboxMessage.save(update_fields=['email_reply', 'is_reply', 'is_read', 'replied_at'])
    return inboxMessage


def get_user_home_message_contact_name(inboxMessage):
    if inboxMessage.message_mode == PlatformInboxMessage.MessageModeChoices.INBOX:
        return inboxMessage.sender_name

    return inboxMessage.recipient_name or inboxMessage.sender_name


def get_user_home_message_contact_email(inboxMessage):
    if inboxMessage.message_mode == PlatformInboxMessage.MessageModeChoices.INBOX:
        return inboxMessage.sender_email

    return inboxMessage.recipient_email or inboxMessage.sender_email


def serialize_user_home_inbox_message(inboxMessage):
    contactName = get_user_home_message_contact_name(inboxMessage)
    contactEmail = get_user_home_message_contact_email(inboxMessage)
    previousMessage = inboxMessage.previous_message or inboxMessage.email_body

    return {
        'id': inboxMessage.id,
        'mode': inboxMessage.message_mode,
        'draftType': inboxMessage.draft_type or None,
        'recipient': contactName or contactEmail,
        'recipientEmail': contactEmail,
        'subject': inboxMessage.email_subject,
        'body': inboxMessage.email_body,
        'previousMessage': previousMessage,
        'replyMessage': (
            inboxMessage.email_body
            if inboxMessage.message_mode == PlatformInboxMessage.MessageModeChoices.DRAFTS
            and inboxMessage.draft_type == PlatformInboxMessage.DraftTypeChoices.REPLY
            else ''
        ),
        'isReply': inboxMessage.is_reply,
        'isRead': inboxMessage.is_read,
    }


def get_user_home_mailbox_state(platformUser):
    ensure_user_home_inbox_messages(platformUser)
    mailboxState = {
        'inbox': [],
        'sent': [],
        'drafts': [],
    }
    inboxMessages = PlatformInboxMessage.objects.filter(
        platform_user=platformUser,
        is_hidden_from_user=False,
    ).order_by('-created_at', '-id')

    for inboxMessage in inboxMessages:
        messageMode = inboxMessage.message_mode
        if messageMode not in mailboxState:
            continue

        mailboxState[messageMode].append(serialize_user_home_inbox_message(inboxMessage))

    return mailboxState


def get_user_home_sender_details(platformUser):
    platformProfile = get_platform_profile(platformUser)

    return {
        'sender_name': platformProfile.full_name if platformProfile else platformUser.username,
        'sender_email': platformUser.email,
    }


def create_user_home_mailbox_message(
    platformUser,
    messageMode,
    draftType,
    recipientValue,
    subjectValue,
    bodyValue,
    previousMessageValue='',
):
    senderDetails = get_user_home_sender_details(platformUser)
    normalizedRecipient = (recipientValue or '').strip()
    normalizedSubject = (subjectValue or '').strip() or 'No subject'
    normalizedBody = (bodyValue or '').strip()
    normalizedPreviousMessage = (previousMessageValue or '').strip()
    recipientEmail = normalizedRecipient if '@' in normalizedRecipient else ''

    return PlatformInboxMessage.objects.create(
        platform_user=platformUser,
        message_mode=messageMode,
        draft_type=draftType or '',
        sender_name=senderDetails['sender_name'],
        sender_email=senderDetails['sender_email'],
        recipient_name=normalizedRecipient,
        recipient_email=recipientEmail,
        email_subject=normalizedSubject,
        email_body=normalizedBody,
        previous_message=normalizedPreviousMessage,
        is_reply=(draftType == PlatformInboxMessage.DraftTypeChoices.REPLY),
        is_read=True,
        replied_at=timezone.now() if draftType == PlatformInboxMessage.DraftTypeChoices.REPLY else None,
    )


def hide_user_home_mailbox_message(platformUser, messageId):
    try:
        normalizedMessageId = int((str(messageId or '')).strip())
    except (TypeError, ValueError):
        return None

    inboxMessage = PlatformInboxMessage.objects.filter(
        platform_user=platformUser,
        pk=normalizedMessageId,
    ).first()
    if inboxMessage is None:
        return None

    hiddenMessageData = serialize_user_home_inbox_message(inboxMessage)
    inboxMessage.is_hidden_from_user = True
    inboxMessage.save(update_fields=['is_hidden_from_user'])
    return hiddenMessageData


def parse_schedule_date(dateValue):
    normalizedDate = (dateValue or '').strip()
    if not normalizedDate:
        return None

    try:
        return datetime.strptime(normalizedDate, '%Y-%m-%d').date()
    except ValueError:
        return None


def parse_schedule_time(timeValue):
    normalizedTime = (timeValue or '').strip()
    if not normalizedTime:
        return None

    try:
        return datetime.strptime(normalizedTime, '%H:%M').time()
    except ValueError:
        return None


def get_schedule_time_minutes(timeValue):
    return timeValue.hour * 60 + timeValue.minute


def get_paired_schedule_color(primaryColor):
    normalizedPrimary = (primaryColor or '').strip().lower()
    for colorIndex, paletteColor in enumerate(SCHEDULE_COLOR_PALETTE):
        if paletteColor.lower() == normalizedPrimary:
            return SCHEDULE_COLOR_PALETTE[(colorIndex + 1) % len(SCHEDULE_COLOR_PALETTE)]

    return 'rgba(99, 102, 241, 1)'


def build_user_home_schedule_event_seed(selectedDate=None):
    selectedDate = selectedDate or timezone.localdate()
    daysInMonth = monthrange(selectedDate.year, selectedDate.month)[1]
    scheduleSeedItems = []

    for dayValue in range(1, daysInMonth + 1):
        eventDate = datetime(selectedDate.year, selectedDate.month, dayValue).date()
        dayTemplates = SCHEDULE_EVENT_WEEKDAY_BLUEPRINT.get(eventDate.isoweekday(), [])

        for dayTemplate in dayTemplates:
            scheduleSeedItems.append(
                {
                    **dayTemplate,
                    'event_date': eventDate,
                }
            )

    return scheduleSeedItems


def ensure_user_home_schedule_events(platformUser, selectedDate=None):
    if PlatformScheduleEvent.objects.filter(platform_user=platformUser).exists():
        return

    scheduleSeedItems = build_user_home_schedule_event_seed(selectedDate)
    PlatformScheduleEvent.objects.bulk_create(
        [
            PlatformScheduleEvent(
                platform_user=platformUser,
                title=scheduleSeedItem['title'],
                event_date=scheduleSeedItem['event_date'],
                start_time=parse_schedule_time(scheduleSeedItem['start_time']),
                end_time=parse_schedule_time(scheduleSeedItem['end_time']),
                platform=scheduleSeedItem['platform'],
                invite_members=scheduleSeedItem['invite_members'],
                color=scheduleSeedItem['color'],
                color_secondary=scheduleSeedItem['color_secondary'],
            )
            for scheduleSeedItem in scheduleSeedItems
        ]
    )


def serialize_user_home_schedule_event(scheduleEvent, detailUrl=''):
    startMinutes = get_schedule_time_minutes(scheduleEvent.start_time)
    endMinutes = get_schedule_time_minutes(scheduleEvent.end_time)
    primaryColor = scheduleEvent.color or 'rgba(37, 99, 235, 1)'
    secondaryColor = scheduleEvent.color_secondary or get_paired_schedule_color(primaryColor)

    return {
        'id': scheduleEvent.id,
        'source': 'db',
        'date': scheduleEvent.event_date.isoformat(),
        'startTime': scheduleEvent.start_time.strftime('%H:%M'),
        'endTime': scheduleEvent.end_time.strftime('%H:%M'),
        'startMinutes': startMinutes,
        'endMinutes': endMinutes,
        'platform': scheduleEvent.platform,
        'inviteMembers': scheduleEvent.invite_members,
        'title': scheduleEvent.title,
        'color': primaryColor,
        'colorSecondary': secondaryColor,
        'detailUrl': detailUrl,
    }


def get_user_home_schedule_event_state(platformUser, selectedDate=None, detailUrl=''):
    ensure_user_home_schedule_events(platformUser, selectedDate)
    scheduleEvents = PlatformScheduleEvent.objects.filter(platform_user=platformUser).order_by(
        'event_date',
        'start_time',
        'id',
    )

    return [
        serialize_user_home_schedule_event(scheduleEvent, detailUrl=detailUrl)
        for scheduleEvent in scheduleEvents
    ]


def create_user_home_schedule_event(platformUser, requestData):
    eventDate = parse_schedule_date(requestData.get('date'))
    startTime = parse_schedule_time(requestData.get('startTime') or requestData.get('start_time'))
    endTime = parse_schedule_time(requestData.get('endTime') or requestData.get('end_time'))

    if eventDate is None:
        return None, 'Choose a valid event date.'

    if startTime is None or endTime is None:
        return None, 'Choose valid start and end times.'

    startMinutes = get_schedule_time_minutes(startTime)
    endMinutes = get_schedule_time_minutes(endTime)

    if endMinutes <= startMinutes:
        return None, 'End time must be later than start time.'

    if startMinutes < SCHEDULE_DAY_START_MINUTES or endMinutes > SCHEDULE_DAY_END_MINUTES:
        return None, 'Meeting time must be between 07:00 and 23:00.'

    platformValue = (requestData.get('platform') or '').strip()
    inviteMembersValue = (requestData.get('inviteMembers') or requestData.get('invite_members') or '').strip()

    if not platformValue:
        return None, 'Choose a platform.'

    if not inviteMembersValue:
        return None, 'Add invite members.'

    colorValue = (requestData.get('color') or '').strip() or 'rgba(37, 99, 235, 1)'
    colorSecondaryValue = (
        (requestData.get('colorSecondary') or requestData.get('color_secondary') or '').strip()
        or get_paired_schedule_color(colorValue)
    )
    titleValue = (requestData.get('title') or '').strip()

    scheduleEvent = PlatformScheduleEvent.objects.create(
        platform_user=platformUser,
        title=titleValue,
        event_date=eventDate,
        start_time=startTime,
        end_time=endTime,
        platform=platformValue,
        invite_members=inviteMembersValue,
        color=colorValue,
        color_secondary=colorSecondaryValue,
    )
    PlatformDashboardActivity.objects.create(
        platform_user=platformUser,
        activity_text=f'Scheduled {titleValue or "an event"} with {inviteMembersValue}',
        activity_icon=PlatformDashboardActivity.ActivityIconChoices.CALENDAR,
    )

    return scheduleEvent, None


def get_user_home_quick_tools_catalog():
    return USER_HOME_QUICK_TOOLS_CATALOG


def get_user_home_quick_tool_slot_state(platformUser):
    slotState = [None] * USER_HOME_QUICK_TOOL_SLOT_COUNT
    platformSlots = PlatformDashboardQuickToolSlot.objects.filter(platform_user=platformUser)

    for platformSlot in platformSlots:
        if (
            platformSlot.slot_index >= USER_HOME_QUICK_TOOL_SLOT_COUNT
            or platformSlot.tool_id not in USER_HOME_QUICK_TOOL_IDS
        ):
            continue

        slotState[platformSlot.slot_index] = platformSlot.tool_id

    return slotState


def validate_user_home_quick_tool_slot_state(slotState):
    if not isinstance(slotState, list):
        return None, 'slot_state must be a list.'

    if len(slotState) != USER_HOME_QUICK_TOOL_SLOT_COUNT:
        return (
            None,
            f'slot_state must contain exactly {USER_HOME_QUICK_TOOL_SLOT_COUNT} slots.',
        )

    normalizedSlotState = []
    for toolId in slotState:
        if toolId is None:
            normalizedSlotState.append(None)
            continue

        if not isinstance(toolId, str):
            return None, 'slot_state entries must be either null or valid tool ids.'

        normalizedToolId = toolId.strip()
        if not normalizedToolId:
            normalizedSlotState.append(None)
            continue

        if normalizedToolId not in USER_HOME_QUICK_TOOL_IDS:
            return None, f'Invalid tool id: {normalizedToolId}'

        normalizedSlotState.append(normalizedToolId)

    filledToolIds = [toolId for toolId in normalizedSlotState if toolId]
    if len(filledToolIds) != len(set(filledToolIds)):
        return None, 'Duplicate quick tools are not allowed.'

    return normalizedSlotState, None


def save_user_home_quick_tool_slot_state(platformUser, slotState):
    normalizedSlotState, validationError = validate_user_home_quick_tool_slot_state(slotState)
    if validationError:
        raise ValueError(validationError)

    platformSlotsByIndex = {
        platformSlot.slot_index: platformSlot
        for platformSlot in PlatformDashboardQuickToolSlot.objects.filter(platform_user=platformUser)
    }

    for slotIndex in range(USER_HOME_QUICK_TOOL_SLOT_COUNT):
        toolId = normalizedSlotState[slotIndex]
        existingSlot = platformSlotsByIndex.get(slotIndex)

        if toolId is None:
            if existingSlot:
                existingSlot.delete()
            continue

        if existingSlot:
            if existingSlot.tool_id != toolId:
                existingSlot.tool_id = toolId
                existingSlot.save(update_fields=['tool_id', 'updated_at'])
            continue

        PlatformDashboardQuickToolSlot.objects.create(
            platform_user=platformUser,
            slot_index=slotIndex,
            tool_id=toolId,
        )

    return normalizedSlotState


def build_user_home_activity_seed(platformUser):
    platformProfile = get_platform_profile(platformUser)
    profileName = platformProfile.full_name if platformProfile else platformUser.username
    teamName = get_platform_profile_team_name(platformProfile) or 'Your team'

    return [
        {
            'activity_text': f'{teamName} added a new block to the main dashboard',
            'activity_icon': PlatformDashboardActivity.ActivityIconChoices.GITHUB,
        },
        {
            'activity_text': f'{profileName} profile was refreshed as a placeholder activity',
            'activity_icon': PlatformDashboardActivity.ActivityIconChoices.TEAM,
        },
        {
            'activity_text': 'Calendar placeholder activity was prepared for future sync updates',
            'activity_icon': PlatformDashboardActivity.ActivityIconChoices.CALENDAR,
        },
        {
            'activity_text': 'Report placeholder activity was prepared for future report generation',
            'activity_icon': PlatformDashboardActivity.ActivityIconChoices.REPORT,
        },
        {
            'activity_text': 'Message placeholder activity was prepared for quick e-mail actions',
            'activity_icon': PlatformDashboardActivity.ActivityIconChoices.MESSAGE,
        },
    ]


def ensure_user_home_activities(platformUser):
    activitySeed = build_user_home_activity_seed(platformUser)
    if not activitySeed:
        return

    allUserActivities = PlatformDashboardActivity.objects.filter(platform_user=platformUser)
    placeholderActivityTexts = [activityItem['activity_text'] for activityItem in activitySeed]
    placeholderActivities = list(
        allUserActivities.filter(activity_text__in=placeholderActivityTexts).order_by('-created_at', '-id')
    )
    hasNonPlaceholderActivities = allUserActivities.exclude(
        activity_text__in=placeholderActivityTexts
    ).exists()

    if allUserActivities.exists():
        if not hasNonPlaceholderActivities and len(placeholderActivities) > 1:
            placeholderActivityIdsToRemove = [activityItem.id for activityItem in placeholderActivities[1:]]
            PlatformDashboardActivity.objects.filter(
                platform_user=platformUser,
                id__in=placeholderActivityIdsToRemove,
            ).delete()
        return

    firstActivityItem = activitySeed[0]
    PlatformDashboardActivity.objects.create(
        platform_user=platformUser,
        activity_text=firstActivityItem['activity_text'],
        activity_icon=firstActivityItem['activity_icon'],
    )


def get_user_home_recent_activities(platformUser, limit=5):
    ensure_user_home_activities(platformUser)
    placeholderSeedTexts = [activityItem['activity_text'] for activityItem in build_user_home_activity_seed(platformUser)]
    activityVisibleSince = timezone.now() - timedelta(hours=USER_HOME_ACTIVITY_VISIBLE_HOURS)

    return list(
        PlatformDashboardActivity.objects.filter(platform_user=platformUser)
        .filter(
            ~Q(activity_text__in=placeholderSeedTexts)
            | Q(
                activity_text__in=placeholderSeedTexts,
                created_at__gte=activityVisibleSince,
            )
        )
        .order_by('-created_at')[:limit]
    )


def get_user_home_db_probe_character(platformUser, recentActivity):
    if recentActivity:
        firstActivityText = (recentActivity[0].activity_text or '').strip()
        if firstActivityText:
            return firstActivityText[0]

    latestActivityText = (
        PlatformDashboardActivity.objects.filter(platform_user=platformUser)
        .order_by('-created_at')
        .values_list('activity_text', flat=True)
        .first()
    )
    if latestActivityText:
        normalizedActivityText = latestActivityText.strip()
        if normalizedActivityText:
            return normalizedActivityText[0]

    platformProfile = get_platform_profile(platformUser)
    if platformProfile and platformProfile.full_name:
        normalizedProfileName = platformProfile.full_name.strip()
        if normalizedProfileName:
            return normalizedProfileName[0]

    normalizedUsername = (platformUser.username or '').strip()
    if normalizedUsername:
        return normalizedUsername[0]

    return '-'


def search_user_home_people(queryValue, limit=8):
    cleanedQuery = (queryValue or '').strip()
    if len(cleanedQuery) < 3:
        return []

    user_model = get_user_model()
    matchingUsersQuerySet = (
        user_model.objects.filter(is_active=True, platform_profile__isnull=False)
        .select_related('platform_profile', 'platform_profile__platform_team')
    )
    userTypeChoices = getattr(user_model, 'UserTypeChoices', None)
    if userTypeChoices and hasattr(userTypeChoices, 'USER'):
        matchingUsersQuerySet = matchingUsersQuerySet.filter(user_type=userTypeChoices.USER)

    matchingUsers = (
        matchingUsersQuerySet.filter(
            Q(username__icontains=cleanedQuery)
            | Q(email__icontains=cleanedQuery)
            | Q(platform_profile__full_name__icontains=cleanedQuery)
            | Q(platform_profile__team_role__icontains=cleanedQuery)
            | Q(platform_profile__team_name__icontains=cleanedQuery)
            | Q(platform_profile__platform_team__name__icontains=cleanedQuery)
            | Q(platform_profile__department_name__icontains=cleanedQuery)
            | Q(platform_profile__department_head__icontains=cleanedQuery)
            | Q(platform_profile__status__icontains=cleanedQuery)
            | Q(platform_profile__member_skills__icontains=cleanedQuery)
        )
        .order_by('username')
        .distinct()[:limit]
    )

    searchResults = []
    for matchedUser in matchingUsers:
        platformProfile = get_platform_profile(matchedUser)
        searchResults.append(
            {
                'full_name': (
                    platformProfile.full_name
                    if platformProfile and platformProfile.full_name
                    else matchedUser.username
                ),
                'team_role': (
                    platformProfile.team_role
                    if platformProfile and platformProfile.team_role
                    else 'Team member'
                ),
                'status': (
                    platformProfile.status
                    if platformProfile and platformProfile.status
                    else 'Offline'
                ),
                'email': matchedUser.email,
                'avatar_url': get_profile_image_data_url(platformProfile),
            }
        )

    return searchResults


def build_user_home_context(platformUser):
    platformProfile = get_platform_profile(platformUser)
    if platformProfile:
        ensure_platform_team_for_profile(platformProfile)

    recentActivity = get_user_home_recent_activities(platformUser)
    notificationMessages = get_user_home_notification_messages(platformUser)
    unreadMessageCount = get_user_home_unread_message_count(platformUser)

    return {
        'platform_profile': platformProfile,
        'profile_avatar_url': get_profile_image_data_url(platformProfile),
        'user_home_quick_tools_catalog': get_user_home_quick_tools_catalog(),
        'user_home_quick_tool_slot_state': get_user_home_quick_tool_slot_state(platformUser),
        'user_home_recent_activity': recentActivity,
        'user_home_notification_messages': notificationMessages,
        'user_home_unread_message_count': unreadMessageCount,
        'user_home_has_unread_messages': unreadMessageCount > 0,
        'user_home_db_probe_character': get_user_home_db_probe_character(platformUser, recentActivity),
    }


# ---------------------------------------------------------------------------
# Schedule event update / delete
# ---------------------------------------------------------------------------

def get_user_home_schedule_event_for_user(platformUser, eventId):
    if not eventId:
        return None
    try:
        eventIdValue = int(eventId)
    except (TypeError, ValueError):
        return None

    return PlatformScheduleEvent.objects.filter(
        platform_user=platformUser,
        id=eventIdValue,
    ).first()


def update_user_home_schedule_event(platformUser, eventId, requestData):
    scheduleEvent = get_user_home_schedule_event_for_user(platformUser, eventId)
    if scheduleEvent is None:
        return None, 'Event was not found.'

    eventDate = parse_schedule_date(requestData.get('date'))
    startTime = parse_schedule_time(requestData.get('startTime') or requestData.get('start_time'))
    endTime = parse_schedule_time(requestData.get('endTime') or requestData.get('end_time'))

    if eventDate is None:
        return None, 'Choose a valid event date.'

    if startTime is None or endTime is None:
        return None, 'Choose valid start and end times.'

    startMinutes = get_schedule_time_minutes(startTime)
    endMinutes = get_schedule_time_minutes(endTime)

    if endMinutes <= startMinutes:
        return None, 'End time must be later than start time.'

    if startMinutes < SCHEDULE_DAY_START_MINUTES or endMinutes > SCHEDULE_DAY_END_MINUTES:
        return None, 'Meeting time must be between 07:00 and 23:00.'

    platformValue = (requestData.get('platform') or '').strip()
    inviteMembersValue = (requestData.get('inviteMembers') or requestData.get('invite_members') or '').strip()

    if not platformValue:
        return None, 'Choose a platform.'

    if not inviteMembersValue:
        return None, 'Add invite members.'

    colorValue = (requestData.get('color') or '').strip() or scheduleEvent.color
    colorSecondaryValue = (
        (requestData.get('colorSecondary') or requestData.get('color_secondary') or '').strip()
        or get_paired_schedule_color(colorValue)
    )
    titleValue = (requestData.get('title') or '').strip()

    scheduleEvent.title = titleValue
    scheduleEvent.event_date = eventDate
    scheduleEvent.start_time = startTime
    scheduleEvent.end_time = endTime
    scheduleEvent.platform = platformValue
    scheduleEvent.invite_members = inviteMembersValue
    scheduleEvent.color = colorValue
    scheduleEvent.color_secondary = colorSecondaryValue
    scheduleEvent.save()

    log_dashboard_activity(
        platformUser,
        f'Updated {titleValue or "an event"} with {inviteMembersValue}',
        PlatformDashboardActivity.ActivityIconChoices.CALENDAR,
    )

    return scheduleEvent, None


def delete_user_home_schedule_event(platformUser, eventId):
    scheduleEvent = get_user_home_schedule_event_for_user(platformUser, eventId)
    if scheduleEvent is None:
        return False, 'Event was not found.'

    eventLabel = scheduleEvent.title or scheduleEvent.invite_members
    scheduleEvent.delete()

    log_dashboard_activity(
        platformUser,
        f'Cancelled {eventLabel}',
        PlatformDashboardActivity.ActivityIconChoices.CALENDAR,
    )

    return True, None


# ---------------------------------------------------------------------------
# Activity logging
# ---------------------------------------------------------------------------

def log_dashboard_activity(platformUser, activityText, activityIcon=None):
    iconValue = activityIcon or PlatformDashboardActivity.ActivityIconChoices.GITHUB
    activityTextValue = (activityText or '').strip()
    if not activityTextValue:
        return None

    return PlatformDashboardActivity.objects.create(
        platform_user=platformUser,
        activity_text=activityTextValue[:255],
        activity_icon=iconValue,
    )


# ---------------------------------------------------------------------------
# Report / chart save and export
# ---------------------------------------------------------------------------

REPORT_DOC_CONTENT_TYPES = {
    PlatformReport.DocTypeChoices.PDF: 'application/pdf',
    PlatformReport.DocTypeChoices.XLSX: (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ),
}


def normalize_report_payload(payload):
    if not isinstance(payload, dict):
        return {}

    titleValue = str(payload.get('title') or 'Report').strip() or 'Report'
    subtitleValue = str(payload.get('subtitle') or '').strip()

    metaItems = []
    rawMeta = payload.get('meta') or []
    if isinstance(rawMeta, list):
        for entry in rawMeta:
            if isinstance(entry, dict):
                label = str(entry.get('label') or '').strip()
                value = str(entry.get('value') or '').strip()
            elif isinstance(entry, (list, tuple)) and len(entry) >= 2:
                label = str(entry[0] or '').strip()
                value = str(entry[1] or '').strip()
            else:
                continue
            if label or value:
                metaItems.append({'label': label, 'value': value})

    sections = []
    rawSections = payload.get('sections') or []
    if isinstance(rawSections, list):
        for sectionEntry in rawSections:
            if not isinstance(sectionEntry, dict):
                continue
            sectionHeading = str(sectionEntry.get('heading') or '').strip()
            sectionParagraph = str(sectionEntry.get('paragraph') or '').strip()
            sectionRows = sectionEntry.get('rows') or []
            normalizedRows = []
            if isinstance(sectionRows, list):
                for rowEntry in sectionRows:
                    if isinstance(rowEntry, (list, tuple)):
                        normalizedRows.append([str(cell) for cell in rowEntry])
            sections.append({
                'heading': sectionHeading,
                'paragraph': sectionParagraph,
                'rows': normalizedRows,
            })

    return {
        'title': titleValue,
        'subtitle': subtitleValue,
        'meta': metaItems,
        'sections': sections,
    }


def render_report_pdf(payload):
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    normalized = normalize_report_payload(payload)
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=normalized['title'],
    )
    styles = getSampleStyleSheet()
    titleStyle = ParagraphStyle(
        'ReportTitle',
        parent=styles['Title'],
        textColor=colors.HexColor('#17345f'),
        spaceAfter=6,
    )
    subtitleStyle = ParagraphStyle(
        'ReportSubtitle',
        parent=styles['Normal'],
        textColor=colors.HexColor('#445b7a'),
        spaceAfter=10,
        fontSize=10,
    )
    metaStyle = ParagraphStyle(
        'ReportMeta',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#536175'),
        spaceAfter=4,
    )
    sectionHeadingStyle = ParagraphStyle(
        'ReportSectionHeading',
        parent=styles['Heading2'],
        textColor=colors.HexColor('#17345f'),
        spaceBefore=12,
        spaceAfter=6,
    )
    bodyStyle = ParagraphStyle(
        'ReportBody',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        spaceAfter=8,
    )

    story = [Paragraph(normalized['title'], titleStyle)]
    if normalized['subtitle']:
        story.append(Paragraph(normalized['subtitle'], subtitleStyle))

    for metaItem in normalized['meta']:
        labelText = metaItem['label']
        valueText = metaItem['value']
        if labelText and valueText:
            story.append(Paragraph(f'<b>{labelText}:</b> {valueText}', metaStyle))
        elif valueText:
            story.append(Paragraph(valueText, metaStyle))

    if normalized['meta']:
        story.append(Spacer(1, 6))

    for sectionItem in normalized['sections']:
        if sectionItem['heading']:
            story.append(Paragraph(sectionItem['heading'], sectionHeadingStyle))
        if sectionItem['paragraph']:
            story.append(Paragraph(sectionItem['paragraph'], bodyStyle))
        if sectionItem['rows']:
            tableData = sectionItem['rows']
            sectionTable = Table(tableData, hAlign='LEFT')
            sectionTable.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#17345f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f6fb')]),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#c5cfdd')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(sectionTable)
            story.append(Spacer(1, 6))

    doc.build(story)
    return output.getvalue()


def render_report_xlsx(payload):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    normalized = normalize_report_payload(payload)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (normalized['title'] or 'Report')[:30]

    headerFill = PatternFill('solid', fgColor='17345F')
    headerFont = Font(color='FFFFFF', bold=True)
    titleFont = Font(bold=True, size=16, color='17345F')
    sectionHeadingFont = Font(bold=True, size=12, color='17345F')

    currentRow = 1
    sheet.cell(row=currentRow, column=1, value=normalized['title']).font = titleFont
    currentRow += 1
    if normalized['subtitle']:
        sheet.cell(row=currentRow, column=1, value=normalized['subtitle'])
        currentRow += 1

    for metaItem in normalized['meta']:
        labelCell = sheet.cell(row=currentRow, column=1, value=metaItem['label'])
        labelCell.font = Font(bold=True)
        sheet.cell(row=currentRow, column=2, value=metaItem['value'])
        currentRow += 1

    if normalized['meta']:
        currentRow += 1

    for sectionItem in normalized['sections']:
        if sectionItem['heading']:
            headingCell = sheet.cell(row=currentRow, column=1, value=sectionItem['heading'])
            headingCell.font = sectionHeadingFont
            currentRow += 1
        if sectionItem['paragraph']:
            sheet.cell(row=currentRow, column=1, value=sectionItem['paragraph']).alignment = Alignment(
                wrap_text=True
            )
            currentRow += 1
        for rowIndex, rowValues in enumerate(sectionItem['rows']):
            for columnIndex, cellValue in enumerate(rowValues, start=1):
                writtenCell = sheet.cell(row=currentRow, column=columnIndex, value=cellValue)
                if rowIndex == 0:
                    writtenCell.fill = headerFill
                    writtenCell.font = headerFont
            currentRow += 1
        currentRow += 1

    for columnIndex in range(1, sheet.max_column + 1):
        maxLength = 0
        columnLetter = sheet.cell(row=1, column=columnIndex).column_letter
        for cell in sheet[columnLetter]:
            valueText = '' if cell.value is None else str(cell.value)
            if len(valueText) > maxLength:
                maxLength = len(valueText)
        sheet.column_dimensions[columnLetter].width = min(60, max(12, maxLength + 2))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def render_chart_xlsx(payload):
    return render_report_xlsx(payload)


def save_platform_report(
    platformUser,
    kind,
    reportType,
    docType,
    title,
    payload,
    fileBytes,
    fileContentType,
    fileName='',
):
    kindValue = kind if kind in PlatformReport.KindChoices.values else PlatformReport.KindChoices.REPORT
    docValue = docType if docType in PlatformReport.DocTypeChoices.values else PlatformReport.DocTypeChoices.PDF
    titleValue = (title or '').strip() or 'Untitled report'
    reportTypeValue = (reportType or '').strip() or 'General'
    payloadValue = normalize_report_payload(payload)

    return PlatformReport.objects.create(
        platform_user=platformUser,
        kind=kindValue,
        report_type=reportTypeValue[:100],
        doc_type=docValue,
        title=titleValue[:255],
        payload_json=payloadValue,
        file_blob=bytes(fileBytes) if fileBytes else None,
        file_content_type=fileContentType[:100] if fileContentType else '',
        file_name=(fileName or '')[:255],
    )


def get_platform_report_for_user(platformUser, reportId):
    if not reportId:
        return None
    try:
        reportIdValue = int(reportId)
    except (TypeError, ValueError):
        return None

    return PlatformReport.objects.filter(
        platform_user=platformUser,
        id=reportIdValue,
    ).first()


def list_platform_reports(platformUser, kind=None):
    queryset = PlatformReport.objects.filter(platform_user=platformUser)
    if kind:
        queryset = queryset.filter(kind=kind)
    return list(queryset)


def build_report_file_name(reportType, docType, title=''):
    safeTitle = slugify(title) or slugify(reportType) or 'report'
    extension = 'pdf' if docType == PlatformReport.DocTypeChoices.PDF else 'xlsx'
    dateLabel = timezone.localdate().isoformat()
    return f'{safeTitle}-{dateLabel}.{extension}'
